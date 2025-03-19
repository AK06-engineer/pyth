
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import re
from datetime import datetime

# Tạo hoặc mở file Excel
excel_file = "dang_ky_hoc_phan.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Mã SV", "Họ tên", "Ngày sinh", "Email", "SĐT", "Học kỳ", "Năm học", "Môn học"])
    wb.save(excel_file)

# Hàm kiểm tra hợp lệ
def validate():
    mssv = entry_mssv.get()
    hoten = entry_hoten.get()
    ngaysinh = entry_ngaysinh.get()
    email = entry_email.get()
    sdt = entry_sdt.get()
    hocky = entry_hocky.get()
    namhoc = var_namhoc.get()

    # Lấy các môn học đã chọn
    monhoc = []
    if var_python.get(): monhoc.append("Lập trình Python")
    if var_java.get(): monhoc.append("Lập trình Java")
    if var_cnpm.get(): monhoc.append("Công nghệ phần mềm")
    if var_web.get(): monhoc.append("Phát triển ứng dụng web")

    # Kiểm tra điều kiện
    if not (mssv.isdigit() and len(mssv) == 7):
        messagebox.showerror("Lỗi", "Mã số sinh viên phải là 7 chữ số.")
        return
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showerror("Lỗi", "Email không hợp lệ.")
        return
    if not (sdt.isdigit() and len(sdt) == 10):
        messagebox.showerror("Lỗi", "Số điện thoại phải là 10 chữ số.")
        return
    if hocky not in ["1", "2", "3"]:
        messagebox.showerror("Lỗi", "Học kỳ chỉ được nhập 1, 2 hoặc 3.")
        return
    try:
        datetime.strptime(ngaysinh, "%d/%m/%Y")
    except:
        messagebox.showerror("Lỗi", "Ngày sinh phải có định dạng dd/mm/yyyy.")
        return
    if not monhoc:
        messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 môn học.")
        return

    # Lưu dữ liệu vào Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    for mh in monhoc:
        ws.append([mssv, hoten, ngaysinh, email, sdt, hocky, namhoc, mh])
    wb.save(excel_file)
    messagebox.showinfo("Thành công", "Đăng ký học phần thành công!")

# Thoát chương trình
def thoat():
    root.destroy()

# Giao diện tkinter
root = tk.Tk()
root.title("Đăng ký học phần")
root.geometry("500x550")
root.configure(bg="#b2f2bb")

tk.Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", font=("Arial", 16, "bold"), bg="#b2f2bb").pack(pady=10)

frame = tk.Frame(root, bg="#b2f2bb")
frame.pack()

tk.Label(frame, text="Mã số sinh viên:", bg="#b2f2bb").grid(row=0, column=0, sticky="w")
entry_mssv = tk.Entry(frame)
entry_mssv.grid(row=0, column=1)

tk.Label(frame, text="Họ tên:", bg="#b2f2bb").grid(row=1, column=0, sticky="w")
entry_hoten = tk.Entry(frame)
entry_hoten.grid(row=1, column=1)

tk.Label(frame, text="Ngày sinh (dd/mm/yyyy):", bg="#b2f2bb").grid(row=2, column=0, sticky="w")
entry_ngaysinh = tk.Entry(frame)
entry_ngaysinh.grid(row=2, column=1)

tk.Label(frame, text="Email:", bg="#b2f2bb").grid(row=3, column=0, sticky="w")
entry_email = tk.Entry(frame)
entry_email.grid(row=3, column=1)

tk.Label(frame, text="Số điện thoại:", bg="#b2f2bb").grid(row=4, column=0, sticky="w")
entry_sdt = tk.Entry(frame)
entry_sdt.grid(row=4, column=1)

tk.Label(frame, text="Học kỳ (1/2/3):", bg="#b2f2bb").grid(row=5, column=0, sticky="w")
entry_hocky = tk.Entry(frame)
entry_hocky.grid(row=5, column=1)

tk.Label(frame, text="Năm học:", bg="#b2f2bb").grid(row=6, column=0, sticky="w")
var_namhoc = tk.StringVar()
dropdown_namhoc = tk.OptionMenu(frame, var_namhoc, "2022-2023", "2023-2024", "2024-2025")
dropdown_namhoc.grid(row=6, column=1)
var_namhoc.set("2023-2024")

tk.Label(frame, text="Chọn môn học:", bg="#b2f2bb").grid(row=7, column=0, sticky="w")
var_python = tk.BooleanVar()
var_java = tk.BooleanVar()
var_cnpm = tk.BooleanVar()
var_web = tk.BooleanVar()
tk.Checkbutton(frame, text="Lập trình Python", variable=var_python, bg="#b2f2bb").grid(row=8, column=0, sticky="w")
tk.Checkbutton(frame, text="Lập trình Java", variable=var_java, bg="#b2f2bb").grid(row=8, column=1, sticky="w")
tk.Checkbutton(frame, text="Công nghệ phần mềm", variable=var_cnpm, bg="#b2f2bb").grid(row=9, column=0, sticky="w")
tk.Checkbutton(frame, text="Phát triển ứng dụng web", variable=var_web, bg="#b2f2bb").grid(row=9, column=1, sticky="w")

tk.Button(root, text="Đăng ký", command=validate, bg="#38d9a9", width=15).pack(pady=10)
tk.Button(root, text="Thoát", command=thoat, bg="#ffa8a8", width=15).pack()

root.mainloop()
